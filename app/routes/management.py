import os
import secrets
import qrcode
import zipfile
from io import BytesIO
from flask import (Blueprint, render_template, flash, redirect, url_for, 
                   request, current_app, send_file, after_this_request)

from flask_login import login_required
from app import db
from app.models.student import Student
from app.models.door import Door
from app.forms import StudentForm, DoorForm, ImportStudentsForm, UploadPhotosForm
from app.decorators import admin_required
from werkzeug.utils import secure_filename
import openpyxl
from app.pdf_generator import generate_single_card_pdf, generate_bulk_cards_pdf
import zipfile

bp = Blueprint('management', __name__)

# --- Rutas de Estudiantes ---

@bp.route('/students')
@login_required
@admin_required
def list_students():
    students = Student.query.all()
    return render_template('management/students.html', students=students, title="Gestión de Estudiantes")

@bp.route('/student/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_student():
    form = StudentForm()
    if form.validate_on_submit():
        filename = None
        if form.photo.data:
            # Guardar la foto
            random_hex = secrets.token_hex(8)
            _, f_ext = os.path.splitext(form.photo.data.filename)
            filename = random_hex + f_ext
            photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos', filename)
            form.photo.data.save(photo_path)

        new_student = Student(
            id=form.id.data,
            name=form.name.data,
            course=form.course.data,
            authorized_to_leave=form.authorized_to_leave.data,
            photo=filename
        )
        db.session.add(new_student)
        db.session.commit()
        flash('Estudiante añadido correctamente.', 'success')
        return redirect(url_for('management.list_students'))
    return render_template('management/student_form.html', form=form, title="Añadir Estudiante")

@bp.route('/student/edit/<string:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_student(id):
    student = Student.query.get_or_404(id)
    form = StudentForm(original_id=student.id)
    if form.validate_on_submit():
        if form.photo.data:
             # Eliminar foto anterior si existe
            if student.photo:
                old_photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos', student.photo)
                if os.path.exists(old_photo_path):
                    os.remove(old_photo_path)
            # Guardar nueva foto
            random_hex = secrets.token_hex(8)
            _, f_ext = os.path.splitext(form.photo.data.filename)
            filename = random_hex + f_ext
            photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos', filename)
            form.photo.data.save(photo_path)
            student.photo = filename

        student.id = form.id.data
        student.name = form.name.data
        student.course = form.course.data
        student.authorized_to_leave = form.authorized_to_leave.data
        db.session.commit()
        flash('Estudiante actualizado correctamente.', 'success')
        return redirect(url_for('management.list_students'))
    elif request.method == 'GET':
        form.id.data = student.id
        form.name.data = student.name
        form.course.data = student.course
        form.authorized_to_leave.data = student.authorized_to_leave
    return render_template('management/student_form.html', form=form, title="Editar Estudiante", student=student)

@bp.route('/student/delete/<string:id>', methods=['POST'])
@login_required
@admin_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    if student.photo:
        photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos', student.photo)
        if os.path.exists(photo_path):
            os.remove(photo_path)
    db.session.delete(student)
    db.session.commit()
    flash('Estudiante eliminado correctamente.', 'success')
    return redirect(url_for('management.list_students'))

# --- Rutas de Puertas ---

@bp.route('/doors')
@login_required
@admin_required
def list_doors():
    doors = Door.query.all()
    return render_template('management/doors.html', doors=doors, title="Gestión de Puertas")

@bp.route('/door/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_door():
    form = DoorForm()
    if form.validate_on_submit():
        new_door = Door(name=form.name.data, status=form.status.data)
        db.session.add(new_door)
        db.session.commit()
        flash('Puerta añadida correctamente.', 'success')
        return redirect(url_for('management.list_doors'))
    return render_template('management/door_form.html', form=form, title="Añadir Puerta")

@bp.route('/door/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_door(id):
    door = Door.query.get_or_404(id)
    form = DoorForm()
    if form.validate_on_submit():
        door.name = form.name.data
        door.status = form.status.data
        db.session.commit()
        flash('Puerta actualizada correctamente.', 'success')
        return redirect(url_for('management.list_doors'))
    elif request.method == 'GET':
        form.name.data = door.name
        form.status.data = door.status.name
    return render_template('management/door_form.html', form=form, title="Editar Puerta")

@bp.route('/door/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_door(id):
    door = Door.query.get_or_404(id)
    db.session.delete(door)
    db.session.commit()
    flash('Puerta eliminada correctamente.', 'success')
    return redirect(url_for('management.list_doors'))

@bp.route('/students/template')
@login_required
@admin_required
def download_template():
    """Genera y descarga una plantilla de Excel para la importación."""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Estudiantes"
    
    # Headers
    headers = ["id", "name", "course", "authorized_to_leave"]
    sheet.append(headers)
    
    # Ejemplo
    example = ["1001", "Juan Pérez García", "8A", "SI"]
    sheet.append(example)

    # Añadir una hoja de instrucciones
    instructions_sheet = workbook.create_sheet(title="Instrucciones")
    instructions_sheet['A1'] = "Instrucciones para la importación"
    instructions_sheet.append([])
    instructions_sheet.append(["Columna", "Descripción", "Ejemplo"])
    instructions_sheet.append(["id", "ID o número de carnet único del estudiante.", "1001"])
    instructions_sheet.append(["name", "Nombre completo del estudiante.", "Ana Sofía Rojas"])
    instructions_sheet.append(["course", "Curso del estudiante.", "9B"])
    instructions_sheet.append(["authorized_to_leave", "Indica si el estudiante puede salir solo. Escribir SI o NO.", "SI"])
    
    workbook.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='plantilla_estudiantes.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/students/import', methods=['GET', 'POST'])
@login_required
@admin_required
def import_students():
    form = ImportStudentsForm()
    if form.validate_on_submit():
        file = form.excel_file.data
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        students_to_add = []
        existing_ids = set([s.id for s in Student.query.all()])
        errors = []
        
        # Omitir la fila de encabezado
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            student_id = str(row[0]).strip() if row[0] else None
            name = row[1].strip() if row[1] else None
            course = row[2].strip() if row[2] else None
            auth_str = str(row[3]).strip().upper() if row[3] else "NO"
            
            if not all([student_id, name, course]):
                errors.append(f"Fila {i}: Faltan datos esenciales (ID, Nombre o Curso).")
                continue
            
            if student_id in existing_ids:
                errors.append(f"Fila {i}: El ID '{student_id}' ya existe en la base de datos.")
                continue

            authorized = True if auth_str == 'SI' else False
            
            new_student = Student(
                id=student_id,
                name=name,
                course=course,
                authorized_to_leave=authorized
            )
            students_to_add.append(new_student)
            existing_ids.add(student_id) # Evitar duplicados en el mismo archivo

        if errors:
            for error in errors:
                flash(error, 'danger')
            return redirect(url_for('management.import_students'))
            
        if students_to_add:
            db.session.bulk_save_objects(students_to_add)
            db.session.commit()
            flash(f'Se han importado {len(students_to_add)} estudiantes exitosamente.', 'success')
        else:
            flash('No se encontraron nuevos estudiantes para importar en el archivo.', 'info')
            
        return redirect(url_for('management.list_students'))
        
    return render_template('management/import_students.html', title="Importar Estudiantes", form=form)


@bp.route('/qrcodes/download/zip')
@login_required
@admin_required
def download_qrs_zip():
    students = Student.query.all()
    if not students:
        flash("No hay estudiantes para generar códigos QR.", "warning")
        return redirect(url_for('management.list_students'))

    temp_dir = current_app.config['TEMP_FOLDER']
    zip_filename = os.path.join(temp_dir, 'codigos_qr_estudiantes.zip')

    with zipfile.ZipFile(zip_filename, 'w') as zf:
        for student in students:
            qr_data = student.qr_code_data
            qr_img = qrcode.make(qr_data)
            
            # Guardar la imagen QR en un buffer en memoria
            img_buffer = BytesIO()
            qr_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Nombre del archivo dentro del ZIP
            filename_in_zip = f"{student.id}_{student.name.replace(' ', '_')}.png"
            zf.writestr(filename_in_zip, img_buffer.getvalue())

    @after_this_request
    def cleanup(response):
        try:
            os.remove(zip_filename)
        except Exception as e:
            current_app.logger.error(f"Error al limpiar el archivo ZIP: {e}")
        return response

    return send_file(
        zip_filename,
        as_attachment=True,
        download_name='codigos_qr_estudiantes.zip',
        mimetype='application/zip'
    )

@bp.route('/student/card/<string:id>')
@login_required
@admin_required
def download_single_card(id):
    student = Student.query.get_or_404(id)
    pdf_buffer = generate_single_card_pdf(student)
    
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f'carnet_{student.id}_{student.name}.pdf',
        mimetype='application/pdf'
    )

@bp.route('/students/cards/download/all')
@login_required
@admin_required
def download_all_cards():
    students = Student.query.order_by(Student.course, Student.name).all()
    if not students:
        flash("No hay estudiantes para generar carnets.", "warning")
        return redirect(url_for('management.list_students'))

    pdf_buffer = generate_bulk_cards_pdf(students)
    
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name='todos_los_carnets.pdf',
        mimetype='application/pdf'
    )

# --- NUEVA RUTA PARA CARGAR FOTOS ---

@bp.route('/students/upload-photos', methods=['GET', 'POST'])
@login_required
@admin_required
def upload_photos():
    form = UploadPhotosForm()
    if form.validate_on_submit():
        zip_file_storage = form.zip_file.data
        photos_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos')
        
        updated_count = 0
        failed_ids = []

        try:
            with zipfile.ZipFile(zip_file_storage, 'r') as zip_ref:
                for filename in zip_ref.namelist():
                    # Ignorar archivos de sistema de macOS y subdirectorios
                    if filename.startswith('__MACOSX/') or filename.endswith('/'):
                        continue
                    
                    # Extraer ID del estudiante del nombre del archivo
                    base_filename = os.path.basename(filename)
                    student_id, file_ext = os.path.splitext(base_filename)
                    
                    if file_ext.lower() not in ['.jpg', '.jpeg', '.png']:
                        failed_ids.append(f"{base_filename} (formato no válido)")
                        continue

                    student = Student.query.get(student_id)
                    if student:
                        # Eliminar foto anterior si existe
                        if student.photo:
                            old_photo_path = os.path.join(photos_folder, student.photo)
                            if os.path.exists(old_photo_path):
                                os.remove(old_photo_path)
                        
                        # Crear un nuevo nombre de archivo seguro
                        random_hex = secrets.token_hex(8)
                        new_filename = random_hex + file_ext.lower()
                        
                        # Extraer y guardar la nueva foto
                        photo_data = zip_ref.read(filename)
                        with open(os.path.join(photos_folder, new_filename), 'wb') as f:
                            f.write(photo_data)
                            
                        # Actualizar la base de datos
                        student.photo = new_filename
                        updated_count += 1
                    else:
                        failed_ids.append(f"{student_id} (estudiante no encontrado)")
            
            if updated_count > 0:
                db.session.commit()
                flash(f'Se actualizaron exitosamente las fotos de {updated_count} estudiantes.', 'success')
            
            if failed_ids:
                flash(f'No se pudieron procesar las siguientes fotos: {", ".join(failed_ids)}', 'danger')

            if updated_count == 0 and not failed_ids:
                flash('El archivo ZIP no contenía fotos válidas para estudiantes existentes.', 'info')

            return redirect(url_for('management.list_students'))

        except zipfile.BadZipFile:
            flash('Error: El archivo subido no es un ZIP válido.', 'danger')
        
    return render_template('management/upload_photos.html', title="Cargar Fotos por Lote", form=form)